import sys
import manejo_cajas as mc
import numpy as np
import os

import six
import packaging
import packaging.version
import packaging.specifiers

from PyQt4 import QtCore, QtGui, uic
from openpyxl import Workbook, load_workbook

from matplotlib.figure import Figure
from matplotlib.backends.backend_qt4agg import (
    FigureCanvasQTAgg as FigureCanvas,
    NavigationToolbar2QT as NavigationToolbar)

qtCreatorFile = "GUI_1.3.1.ui"  # Enter file here.

Ui_MainWindow, QtBaseClass = uic.loadUiType(qtCreatorFile)


class MyApp(QtGui.QMainWindow, Ui_MainWindow):
    def __init__(self):
        """
        ESTA FUNCION GENERA UN BUCLE
        INFINITO, PARA PODER VER BIEN
        LA VENTANA.
        """

        QtGui.QMainWindow.__init__(self)
        Ui_MainWindow.__init__(self)
        self.setupUi(self)

        #########DE ACA PARA ABAJO SE AGREGAN LAS FUNCIONES.#########

        ####SELECCIONO DIRECTORIO DE TRABAJO

        self._selecciono_directorio_saves_()

        ####CREACION DE LA MATRIZ DE TRABAJO

        fila_maxima = self.caja_26.maximum()
        columna_maxima = 21

        self.x = np.zeros((fila_maxima, columna_maxima))  # parametros
        self.y = np.zeros((1, 14))  # calibracion
        self.z = np.zeros((1, 11))  # servicios

        self.btn_1.clicked.connect(self._grabar_)
        self.btn_2.clicked.connect(self._cargar_)

        ####OCULTO COSAS

        self._oculto_cosas_()

        self.caja_4.valueChanged.connect(self._oculto_cosas_)
        self.caja_10.valueChanged.connect(self._oculto_cosas_)
        self.caja_13.valueChanged.connect(self._oculto_cosas_)
        self.caja_27.valueChanged.connect(self._oculto_cosas_)
        self.caja_29.valueChanged.connect(self._oculto_cosas_)

        self.checkBox_1.stateChanged.connect(self._oculto_cosas_)
        self.comboBox_2.activated.connect(self._oculto_cosas_)

        ####SETEO MATRIZ

        self.caja_1.valueChanged.connect(self._seteo_matriz_)
        self.caja_2.valueChanged.connect(self._seteo_matriz_)
        self.caja_3.valueChanged.connect(self._seteo_matriz_)
        self.caja_4.valueChanged.connect(self._seteo_matriz_)
        self.caja_5.valueChanged.connect(self._seteo_matriz_)
        self.caja_6.valueChanged.connect(self._seteo_matriz_)
        self.caja_7.valueChanged.connect(self._seteo_matriz_)
        self.caja_8.valueChanged.connect(self._seteo_matriz_)
        self.caja_9.valueChanged.connect(self._seteo_matriz_)
        self.caja_10.valueChanged.connect(self._seteo_matriz_)
        self.caja_11.valueChanged.connect(self._seteo_matriz_)
        self.caja_12.valueChanged.connect(self._seteo_matriz_)
        self.caja_27.valueChanged.connect(self._seteo_matriz_)
        self.caja_28.valueChanged.connect(self._seteo_matriz_)
        self.caja_29.valueChanged.connect(self._seteo_matriz_)
        self.caja_30.valueChanged.connect(self._seteo_matriz_)
        self.caja_13.valueChanged.connect(self._seteo_matriz_)
        self.caja_14.valueChanged.connect(self._seteo_matriz_)
        self.caja_15.valueChanged.connect(self._seteo_matriz_)
        self.caja_16.valueChanged.connect(self._seteo_matriz_)
        self.caja_18.valueChanged.connect(self._seteo_matriz_)
        self.caja_25.valueChanged.connect(self._seteo_matriz_)

        self.caja_20.valueChanged.connect(self._seteo_matriz_)
        self.caja_21.valueChanged.connect(self._seteo_matriz_)
        self.caja_23.valueChanged.connect(self._seteo_matriz_)
        self.caja_24.valueChanged.connect(self._seteo_matriz_)
        self.caja_19.valueChanged.connect(self._seteo_matriz_)
        self.caja_31.valueChanged.connect(self._seteo_matriz_)
        self.caja_22.valueChanged.connect(self._seteo_matriz_)
        self.caja_32.valueChanged.connect(self._seteo_matriz_)
        self.caja_17.valueChanged.connect(self._seteo_matriz_)
        self.caja_33.valueChanged.connect(self._seteo_matriz_)
        self.caja_34.valueChanged.connect(self._seteo_matriz_)
        self.caja_35.valueChanged.connect(self._seteo_matriz_)
        self.caja_36.valueChanged.connect(self._seteo_matriz_)

        self.caja_39.valueChanged.connect(self._seteo_matriz_)
        self.caja_40.valueChanged.connect(self._seteo_matriz_)
        self.caja_41.valueChanged.connect(self._seteo_matriz_)
        self.caja_42.valueChanged.connect(self._seteo_matriz_)
        self.caja_46.valueChanged.connect(self._seteo_matriz_)
        self.caja_47.valueChanged.connect(self._seteo_matriz_)
        self.caja_43.valueChanged.connect(self._seteo_matriz_)
        self.caja_44.valueChanged.connect(self._seteo_matriz_)
        self.caja_45.valueChanged.connect(self._seteo_matriz_)
        self.caja_38.valueChanged.connect(self._seteo_matriz_)
        self.caja_37.valueChanged.connect(self._seteo_matriz_)

        ####SETEO CAJAS

        self.caja_26.valueChanged.connect(self._seteo_cajas_parametros_)

        ####SETEO EXTREMOS

        self.caja_20.valueChanged.connect(self._seteo_extremos_fuerza_)
        self.caja_21.valueChanged.connect(self._seteo_extremos_fuerza_)
        self.caja_23.valueChanged.connect(self._seteo_extremos_fuerza_)
        self.caja_24.valueChanged.connect(self._seteo_extremos_fuerza_)
        self.caja_19.valueChanged.connect(self._seteo_extremos_fuerza_)
        self.caja_31.valueChanged.connect(self._seteo_extremos_fuerza_)
        self.caja_22.valueChanged.connect(self._seteo_extremos_fuerza_)
        self.caja_32.valueChanged.connect(self._seteo_extremos_fuerza_)

        self.caja_20.valueChanged.connect(self._seteo_extremos_corriente_)
        self.caja_21.valueChanged.connect(self._seteo_extremos_corriente_)
        self.caja_23.valueChanged.connect(self._seteo_extremos_corriente_)
        self.caja_24.valueChanged.connect(self._seteo_extremos_corriente_)
        self.caja_19.valueChanged.connect(self._seteo_extremos_corriente_)
        self.caja_31.valueChanged.connect(self._seteo_extremos_corriente_)
        self.caja_22.valueChanged.connect(self._seteo_extremos_corriente_)
        self.caja_32.valueChanged.connect(self._seteo_extremos_corriente_)

        ####CALCULO TIEMPO PROCESO

        self._calculo_tiempo_proceso()

        self.caja_1.valueChanged.connect(self._calculo_tiempo_proceso)
        self.caja_2.valueChanged.connect(self._calculo_tiempo_proceso)
        self.caja_3.valueChanged.connect(self._calculo_tiempo_proceso)
        self.caja_4.valueChanged.connect(self._calculo_tiempo_proceso)
        self.caja_6.valueChanged.connect(self._calculo_tiempo_proceso)
        self.caja_27.valueChanged.connect(self._calculo_tiempo_proceso)
        self.caja_29.valueChanged.connect(self._calculo_tiempo_proceso)
        self.caja_7.valueChanged.connect(self._calculo_tiempo_proceso)
        self.caja_9.valueChanged.connect(self._calculo_tiempo_proceso)
        self.caja_10.valueChanged.connect(self._calculo_tiempo_proceso)
        self.caja_13.valueChanged.connect(self._calculo_tiempo_proceso)
        self.caja_15.valueChanged.connect(self._calculo_tiempo_proceso)
        self.caja_16.valueChanged.connect(self._calculo_tiempo_proceso)

        ####TOOL BAR

        self.actionAbrir.triggered.connect(self. _abro_dat_)
        self.actionGuardar.triggered.connect(self._guardo_en_txt_)
        self.actionGuardar_Como.triggered.connect(self._guardo_como_en_txt_)
        self.actionImportar.triggered.connect(self._cargo_excel_)
        self.actionExportar.triggered.connect(self._grabo_excel_)

        ####MENU

        self.actionNuevo.triggered.connect(self._nuevo_)
        self.actionClonar_Programas.triggered.connect(self._copiar_programas_)
        self.actionConfiguracion.triggered.connect(self._configuracion_)
        self.actionManual.triggered.connect(self._manual_)
        self.actionAcerca_de.triggered.connect(self._acerca_de_)
        self.actionAcerca_de_QT.triggered.connect(self._acerca_de_QT_)

        ####MEDIR CALIBRACION

        self.Medir_Fuerza_1.clicked.connect(lambda: mc.grabo_dato_calibracion(1))
        self.Medir_Fuerza_2.clicked.connect(lambda: mc.grabo_dato_calibracion(2))

        self.Medir_Corriente_1.clicked.connect(lambda: mc.grabo_dato_calibracion(3))
        self.Medir_Corriente_2.clicked.connect(lambda: mc.grabo_dato_calibracion(4))

        #es necesaria la funcion lambda.

    def _grabar_(self):
        """
        ESTA FUNCION TOMA LOS VALORES NUMERICOS
        DE LAS CAJAS QUE HAY EN LA GUI, Y LOS
        CONVIERTE EN UN STRING.
        """

        ###################################################################1%
        self.progressBar.setValue(1)

        Acercamiento_1 = str(self.caja_1.value())
        Acercamiento_2 = str(self.caja_2.value())

        Tiempo_Repe = str(self.caja_3.value())

        Pre_Sold = str(self.caja_4.value())
        Intensidad_1 = str(self._calculo_duty_corriente_(self.caja_5.value()))
        Tiempo_Frio_1 = str(self.caja_6.value())

        Soldadura = str(self.caja_7.value())
        Intensidad_2 = str(self._calculo_duty_corriente_(self.caja_8.value()))
        Tiempo_Frio_2 = str(self.caja_9.value())
        Impulsos = str(self.caja_10.value())

        Post_Sold = str(self.caja_13.value())
        Intensidad_3 = str(self._calculo_duty_corriente_(self.caja_14.value()))
        Tiempo_Frio_3 = str(self.caja_15.value())

        Fuerza = str(self._calculo_duty_fuerza_(self.caja_18.value()))
        Tiempo_Forja = str(self.caja_16.value())

        Incremento = str(self.caja_27.value())
        Decenso = str(self.caja_29.value())
        Incremento_Potencia = str(self.caja_28.value())
        Decenso_Potencia = str(self.caja_30.value())

        Tolerancia_Sup = str(self.caja_11.value())
        Tolerancia_Inf = str(self.caja_12.value())

        Programa = self.caja_26.value()

        mc.grabo_dato_parametro(Programa, Acercamiento_1, Acercamiento_2, Tiempo_Repe)
        mc.grabo_dato_parametro_pre_sold(Programa, Pre_Sold, Intensidad_1, Tiempo_Frio_1)
        mc.grabo_dato_parametro_sold(Programa, Soldadura, Intensidad_2, Tiempo_Frio_2, Impulsos)
        mc.grabo_dato_parametro_post_sold(Programa, Post_Sold, Intensidad_3, Tiempo_Frio_3)
        mc.grabo_dato_parametro_forja(Programa, Tiempo_Forja, Fuerza)
        mc.grabo_dato_parametro_incremento_decenso(Programa, Incremento, Decenso, Incremento_Potencia, Decenso_Potencia)
        mc.grabo_dato_parametro_tolerancia(Programa, Tolerancia_Sup, Tolerancia_Inf)

        ###################################################################33%
        self.progressBar.setValue(33)

        Acercamiento_Calibracion = self.caja_17.value()
        Apriete_Calibracion = self.caja_33.value()

        Presion_Calibracion = str(self.caja_34.value())
        Soldadura_Calibracion = str(self.caja_35.value())
        Mantenido_Calibracion = str(self.caja_36.value())

        Fuerza_Calibracion_1 = str(self.caja_20.value() * 10)
        Fuerza_Calibracion_2 = str(self.caja_21.value() * 10)

        Corriente_Calibracion_1 = str(self.caja_19.value())
        Corriente_Calibracion_2 = str(self.caja_31.value())

        mc.grabo_dato_calibracion_parametros(Acercamiento_Calibracion, Apriete_Calibracion, Presion_Calibracion, Soldadura_Calibracion, Mantenido_Calibracion, Fuerza_Calibracion_1, Fuerza_Calibracion_2, Corriente_Calibracion_1, Corriente_Calibracion_2)

        ###################################################################66%
        self.progressBar.setValue(66)

        Puntos = self.caja_39.value()

        Fresados = str(self.caja_40.value())

        Alarma_Puntos = str(self.caja_41.value())
        Alarma_Fresados = str(self.caja_42.value())

        Fuerza_Fresado = str(self._calculo_duty_fuerza_(self.caja_46.value()))
        Fuerza_Cambio_Electrodo = str(self._calculo_duty_fuerza_(self.caja_47.value()))

        Tiempo_Acercamiento = str(self.caja_43.value())
        Tiempo_Apriete = str(self.caja_44.value())

        Puntos_Pieza = self.caja_45.value()

        Numero_Curva = str(self.caja_38.value())
        Porcentaje = str(self.caja_37.value())

        mc.grabo_dato_servicios_puntos(Puntos, Alarma_Puntos, Fresados, Alarma_Fresados)
        mc.grabo_dato_servicios_fuerza(Fuerza_Fresado, Fuerza_Cambio_Electrodo)
        mc.grabo_dato_servicios_tiempo(Tiempo_Acercamiento, Tiempo_Apriete)
        mc.grabo_dato_servicios_Punto_Piezas(Puntos_Pieza)
        mc.grabo_dato_servicios_Ley_Incremental(Numero_Curva, Porcentaje)

        ###################################################################100%
        self.progressBar.setValue(100)

    def _cargar_(self):
        """
        ESTA FUNCION TOMA LOS VALORES YA GUARDADOS
        EN EL CONTROL DE SOLDADURA Y LO CARGA
        EN LA MATRIZ.
        """

        fila_maxima = self.caja_26.maximum()

        ###################################################################1%
        self.progressBar.setValue(1)
        self.y = mc.pido_dato_eeprom_calibracion()

        ###################################################################33%
        self.progressBar.setValue(33)
        self.z = mc.pido_dato_eeprom_servicios()

        ###################################################################66%
        self.progressBar.setValue(66)
        self.x = mc.pido_dato_eeprom_parametros(fila_maxima)

        ###################################################################100%
        self.progressBar.setValue(100)

        self._seteo_cajas_parametros_()
        self._seteo_cajas_calibracion_()
        self._seteo_cajas_servicios_()
        self._seteo_cajas_configuracion_()

        self._guardo_backup_()

    def _seteo_extremos_fuerza_(self):
        """
        ESTA FUNCION TOMA UN VALOR DE REFERENCIA
        Y LE APLICA A LAS CAJAS UN VALOR MAXIMO
        Y OTRO VALOR MINIMO. PARA LA FUERZA.
        """

        punto_x_1 = self.caja_20.value() * 10
        punto_y_1 = self.caja_23.value()

        punto_x_2 = self.caja_21.value() * 10
        punto_y_2 = self.caja_24.value()

        try:
            valor_maximo = ( (punto_y_2 * punto_x_1) - (punto_y_1 * punto_x_2) + ( (punto_y_1 -  punto_y_2) * 99 ) ) / (punto_x_1 - punto_x_2)
            self.caja_18.setRange(0, valor_maximo)

        except ZeroDivisionError:
            pass

    def _seteo_extremos_corriente_(self):
        """
        ESTA FUNCION TOMA UN VALOR DE REFERENCIA
        Y LE APLICA A LAS CAJAS UN VALOR MAXIMO
        Y OTRO VALOR MINIMO. PARA CORRIENTE.
        """

        punto_x_1 = self.caja_19.value()
        punto_y_1 = self.caja_22.value()

        punto_x_2 = self.caja_31.value()
        punto_y_2 = self.caja_32.value()

        try:
            # valor_minimo = ( (punto_y_2 * punto_x_1) - (punto_y_1 * punto_x_2) + ( (punto_y_1 - punto_y_2) * 1 ) ) / (punto_x_1 - punto_x_2)
            valor_maximo = ( (punto_y_2 * punto_x_1) - (punto_y_1 * punto_x_2) + ( (punto_y_1 - punto_y_2) * 99 ) ) / (punto_x_1 - punto_x_2)

            #if valor_minimo < 1:
            #    valor_minimo = 1

            # self.caja_5.setRange(valor_minimo, valor_maximo)
            # self.caja_8.setRange(valor_minimo, valor_maximo)
            # self.caja_14.setRange(valor_minimo, valor_maximo)

            self.caja_5.setRange(0, valor_maximo)
            self.caja_8.setRange(0, valor_maximo)
            self.caja_14.setRange(0, valor_maximo)

        except ZeroDivisionError:
            pass

    def _calculo_tiempo_proceso(self):
        """
        ESTA FUNCION CALCULA EL TIEMPO
        QUE DURA EL PROCESO DE SOLDADURA.
        """

        tiempo_1 = self.caja_1.value()
        tiempo_2 = self.caja_2.value()
        tiempo_3 = self.caja_3.value()
        tiempo_4 = self.caja_4.value()
        tiempo_5 = self.caja_6.value()
        tiempo_6 = self.caja_27.value()
        tiempo_7 = self.caja_7.value()
        tiempo_8 = (self.caja_9.value() * self.caja_10.value())
        tiempo_9 = self.caja_13.value()

        tiempo_10 = self.caja_15.value()
        tiempo_11 = self.caja_29.value()
        tiempo_12 = self.caja_16.value()

        calculo_1 = tiempo_1 + tiempo_2 + tiempo_3 + tiempo_4 + tiempo_5
        calculo_2 = tiempo_6 + tiempo_7 + tiempo_8 + tiempo_9 + tiempo_10
        calculo_3 = tiempo_11 + tiempo_12

        calculo_4 = (calculo_1 + calculo_2 + calculo_3) * 20

        self.tiempo_proceso.setValue(calculo_4)

    def _calculo_duty_fuerza_(self, fuerza):
        """
        ESTA FUNCION ME CALCULA EL DUTY
        CORRESPONDIENTE DE LA INTENSIDAD
        1/2/3.
        """

        try:
            punto_x_1 = self.caja_20.value() * 10
            punto_y_1 = self.caja_23.value()

            punto_x_2 = self.caja_21.value() * 10
            punto_y_2 = self.caja_24.value()

            return abs(int( ( ( (punto_y_2-fuerza)*punto_x_1 ) - ( (punto_y_1-fuerza)*punto_x_2 ) ) / (punto_y_1 - punto_y_2)))

        except ZeroDivisionError:
            self.caja_23.setValue(1)
            self.caja_24.setValue(2)

            self.caja_22.setValue(1)
            self.caja_32.setValue(2)

    def _calculo_duty_corriente_(self, intensidad):
        """
        ESTA FUNCION ME CALCULA EL DUTY
        CORRESPONDIENTE DE LA INTENSIDAD
        1/2/3.
        """

        try:
            punto_x_1 = self.caja_19.value()
            punto_y_1 = self.caja_22.value()

            punto_x_2 = self.caja_31.value()
            punto_y_2 = self.caja_32.value()

            return abs(int( ( ( (punto_y_2-intensidad)*punto_x_1 ) - ( (punto_y_1-intensidad)*punto_x_2 ) ) / (punto_y_1 - punto_y_2)))

        except ZeroDivisionError:
            self.caja_23.setValue(1)
            self.caja_24.setValue(2)

            self.caja_22.setValue(1)
            self.caja_32.setValue(2)

    def _calculo_valor_fuerza(self, duty):
        """
        """

        x1 = self.caja_20.value() * 10
        y1 = self.caja_23.value()

        x2 = self.caja_21.value() * 10
        y2 = self.caja_24.value()

        # x = ( ( (y - y1) / (y2 - y1) ) * (x2 - x1) ) + (x1)

        return abs(int(((duty - y1) / (y2 - y1)) * (x2 - x1)) + (x1))

    def _calculo_valor_corriente(self, duty):
        """
        """

        x1 = self.caja_19.value()
        y1 = self.caja_22.value()

        x2 = self.caja_31.value()
        y2 = self.caja_32.value()

        return abs(int(((duty - y1) / (y2 - y1)) * (x2 - x1)) + (x1))

    def _oculto_cosas_(self):
        """
        OCULTO LAS CAJAS SEGUN NO SEAN NECESARIAS.
        """

        #REPETICION.
        if(self.comboBox_2.currentText() == "NO"):
            self.caja_3.hide()
            self.label_12.hide()
            self.label_10.hide()
        elif(self.comboBox_2.currentText() == "SI"):
            self.caja_3.show()
            self.label_12.show()
            self.label_10.show()

        #PRE-SOLDADURA.
        if(self.caja_4.value() == 0):
            self.caja_5.hide()
            self.caja_6.hide()
            self.label_28.hide()
            self.label_36.hide()
            self.label_31.hide()
            self.label_37.hide()
        else:
            self.caja_5.show()
            self.caja_6.show()
            self.label_28.show()
            self.label_36.show()
            self.label_31.show()
            self.label_37.show()

        #SOLDADURA.
        if(self.caja_10.value() == 1):
            self.caja_9.hide()
            self.label_32.hide()
            self.label_38.hide()
        else:
            self.caja_9.show()
            self.label_32.show()
            self.label_38.show()

        #POST-SOLDADURA.
        if(self.caja_13.value() == 0):
            self.caja_14.hide()
            self.caja_15.hide()
            self.label_30.hide()
            self.label_42.hide()
            self.label_33.hide()
            self.label_43.hide()
        else:
            self.caja_14.show()
            self.caja_15.show()
            self.label_30.show()
            self.label_42.show()
            self.label_33.show()
            self.label_43.show()

        #REGULACION.
        if(self.checkBox_1.checkState() == 0):
            self.caja_11.hide()
            self.caja_12.hide()
            self.label_18.hide()
            self.label_19.hide()
            self.label_17.hide()
            self.label_20.hide()
        else:
            self.caja_11.show()
            self.caja_12.show()
            self.label_18.show()
            self.label_19.show()
            self.label_17.show()
            self.label_20.show()

        #ASCENSO.
        if(self.caja_27.value() == 0):
            self.caja_28.hide()
            self.label_47.hide()
            self.label_49.hide()
        else:
            self.caja_28.show()
            self.label_47.show()
            self.label_49.show()
        #DESCENSO.
        if(self.caja_29.value() == 0):
            self.caja_30.hide()
            self.label_48.hide()
            self.label_50.hide()
        else:
            self.caja_30.show()
            self.label_48.show()
            self.label_50.show()

    def _seteo_matriz_(self):
        """
        ESTA FUNCION TOMA LOS VALORES ACTUALES DE LAS
        CAJAS Y LOS GUARDA EN UNA MATRIZ.
        """

        fila_actual = self.caja_26.value() - 1

        #////PARAMETROS

        #el problema esta aca :(

        self.x[fila_actual][0] = self.caja_1.value()
        self.x[fila_actual][1] = self.caja_2.value()
        self.x[fila_actual][2] = self.caja_3.value()

        self.x[fila_actual][3] = self.caja_4.value()
        self.x[fila_actual][4] = self.caja_5.value()
        self.x[fila_actual][5] = self.caja_6.value()

        self.x[fila_actual][6] = self.caja_7.value()
        self.x[fila_actual][7] = self.caja_8.value()
        self.x[fila_actual][8] = self.caja_9.value()
        self.x[fila_actual][9] = self.caja_10.value()

        self.x[fila_actual][10] = self.caja_27.value()
        self.x[fila_actual][11] = self.caja_28.value()
        self.x[fila_actual][12] = self.caja_29.value()
        self.x[fila_actual][13] = self.caja_30.value()

        self.x[fila_actual][14] = self.caja_11.value()
        self.x[fila_actual][15] = self.caja_12.value()

        self.x[fila_actual][16] = self.caja_13.value()
        self.x[fila_actual][17] = self.caja_14.value()
        self.x[fila_actual][18] = self.caja_15.value()

        self.x[fila_actual][19] = self.caja_16.value()
        self.x[fila_actual][20] = self.caja_18.value()

        #////CALIBRACION

        self.y[0][0] = self.caja_17.value()
        self.y[0][1] = self.caja_33.value()
        self.y[0][2] = self.caja_34.value()
        self.y[0][3] = self.caja_35.value()
        self.y[0][4] = self.caja_36.value()

        self.y[0][5] = self.caja_20.value()
        self.y[0][6] = self.caja_21.value()
        self.y[0][7] = self.caja_23.value()
        self.y[0][8] = self.caja_24.value()

        self.y[0][9] = self.caja_19.value()
        self.y[0][10] = self.caja_31.value()
        self.y[0][11] = self.caja_22.value()
        self.y[0][12] = self.caja_32.value()

        #////SERVICIOS

        self.z[0][0] = self.caja_39.value()
        self.z[0][1] = self.caja_40.value()
        self.z[0][2] = self.caja_41.value()
        self.z[0][3] = self.caja_42.value()

        self.z[0][4] = self.caja_46.value()
        self.z[0][5] = self.caja_47.value()

        self.z[0][6] = self.caja_43.value()
        self.z[0][7] = self.caja_44.value()
        self.z[0][8] = self.caja_45.value()

        self.z[0][9] = self.caja_38.value()
        self.z[0][10] = self.caja_37.value()

    def _grabo_excel_(self):
        """
        ESTA FUNCION TOMA LOS VALORES DESDE LA MATRIZ Y
        LOS ALMACENA EN UN ARCHIVO EXCEL.
        """

        import datetime
        fecha = str(datetime.datetime.now().date())
        nombre_archivo = "save_" + fecha + ".xlsx"

        direccion = QtGui.QFileDialog.getSaveFileName(self,
                                                      "Abrir",
                                                      nombre_archivo,
                                                      "*.xls *.xlsx"
                                                     )

        wb = Workbook()

        ################################################################HOJA 1
        hoja_parametros_soldadura = wb.active
        hoja_parametros_soldadura.title = "Soldadura"

        ##########ETIQUETAS
        hoja_parametros_soldadura['A1'] = "PROGRAMA/PARAMETRO"

        hoja_parametros_soldadura['B1'] = "ACERCAMIENTO"
        hoja_parametros_soldadura['C1'] = "APRIETE"
        hoja_parametros_soldadura['D1'] = "REPETICION"

        hoja_parametros_soldadura['E1'] = "PRE-SOLDADURA"
        hoja_parametros_soldadura['F1'] = "INTENSIDAD 1"
        hoja_parametros_soldadura['G1'] = "TIEMPO FRIO 1"

        hoja_parametros_soldadura['H1'] = "SOLDADURA"
        hoja_parametros_soldadura['I1'] = "INTENSIDAD 2"
        hoja_parametros_soldadura['J1'] = "TIEMPO FRIO 2"
        hoja_parametros_soldadura['K1'] = "IMPULSOS"

        hoja_parametros_soldadura['L1'] = "ASCENSO"
        hoja_parametros_soldadura['M1'] = "POTENCIA INICIAL"
        hoja_parametros_soldadura['N1'] = "DESCENSO"
        hoja_parametros_soldadura['O1'] = "POTENCIA FINAL"

        hoja_parametros_soldadura['P1'] = "POST-SOLDADURA"
        hoja_parametros_soldadura['Q1'] = "INTENSIDAD 3"
        hoja_parametros_soldadura['R1'] = "TIEMPO FRIO 3"

        hoja_parametros_soldadura['S1'] = "TOLERANCIA SUP"
        hoja_parametros_soldadura['T1'] = "TOLERANCIA INF"

        hoja_parametros_soldadura['U1'] = "TIEMPO FORJA"
        hoja_parametros_soldadura['V1'] = "FUERZA"

        # agregar un for para que tome los valores de la matriz
        fila_maxima = self.caja_26.maximum()
        for i in range(0, fila_maxima):
            Acercamiento_1 = self.x[i][0]
            Acercamiento_2 = self.x[i][1]

            Tiempo_Repe = self.x[i][2]

            Pre_Sold = self.x[i][3]
            Intensidad_1 = self.x[i][4]
            Tiempo_Frio_1 = self.x[i][5]

            Soldadura = self.x[i][6]
            Intensidad_2 = self.x[i][7]
            Tiempo_Frio_2 = self.x[i][8]
            Impulsos = self.x[i][9]

            Post_Sold = self.x[i][10]
            Intensidad_3 = self.x[i][11]
            Tiempo_Frio_3 = self.x[i][12]

            Fuerza = self.x[i][13]
            Tiempo_Forja = self.x[i][14]

            Incremento = self.x[i][15]
            Decenso = self.x[i][16]
            Incremento_Potencia = self.x[i][17]
            Decenso_Potencia = self.x[i][18]

            Tolerancia_Sup = self.x[i][19]
            Tolerancia_Inf = self.x[i][20]

            hoja_parametros_soldadura.append(
                                            [i+1, Acercamiento_1, Acercamiento_2,
                                            Tiempo_Repe,
                                            Pre_Sold, Intensidad_1, Tiempo_Frio_1,
                                            Soldadura, Intensidad_2, Tiempo_Frio_2,
                                            Impulsos,
                                            Post_Sold, Intensidad_3, Tiempo_Frio_3,
                                            Fuerza, Tiempo_Forja,
                                            Incremento, Decenso,
                                            Incremento_Potencia, Decenso_Potencia,
                                            Tolerancia_Sup, Tolerancia_Inf
                                            ]
                                            )

        ################################################################HOJA 2
        hoja_parametros_calibracion = wb.create_sheet(title="Calibracion")

        ##########ETIQUETAS
        hoja_parametros_calibracion['A1'] = "ACERCAMIENTO"
        hoja_parametros_calibracion['B1'] = "APRIETE"
        hoja_parametros_calibracion['C1'] = "PRESION"
        hoja_parametros_calibracion['D1'] = "SOLDADURA"
        hoja_parametros_calibracion['E1'] = "MANTENIDO"

        hoja_parametros_calibracion['F1'] = "% FUERZA 1"
        hoja_parametros_calibracion['G1'] = "VALOR FUERZA 1"
        hoja_parametros_calibracion['H1'] = "% FUERZA 2"
        hoja_parametros_calibracion['I1'] = "VALOR FUERZA 2"

        hoja_parametros_calibracion['J1'] = "% CORRIENTE 1"
        hoja_parametros_calibracion['K1'] = "VALOR CORRIENTE 1"
        hoja_parametros_calibracion['L1'] = "% CORRIENTE 2"
        hoja_parametros_calibracion['M1'] = "VALOR CORRIENTE 2"

        Acercamiento_Calibracion = self.caja_17.value()
        Apriete_Calibracion = self.caja_33.value()
        Presion_Calibracion = self.caja_34.value()
        Soldadura_Calibracion = self.caja_35.value()
        Mantenido_Calibracion = self.caja_36.value()

        Fuerza_Calibracion_1 = self.caja_20.value()
        val_Fuerza_Calibracion_1 = self.caja_23.value()

        Fuerza_Calibracion_2 = self.caja_21.value()
        val_Fuerza_Calibracion_2 = self.caja_24.value()

        Corriente_Calibracion_1 = self.caja_19.value()
        val_Corriente_Calibracion_1 = self.caja_22.value()

        Corriente_Calibracion_2 = self.caja_31.value()
        val_Corriente_Calibracion_2 = self.caja_32.value()

        hoja_parametros_calibracion.append(
                                          [Acercamiento_Calibracion,
                                          Apriete_Calibracion,
                                          Presion_Calibracion,
                                          Soldadura_Calibracion,
                                          Mantenido_Calibracion,
                                          Fuerza_Calibracion_1,
                                          val_Fuerza_Calibracion_1,
                                          Fuerza_Calibracion_2,
                                          val_Fuerza_Calibracion_2,
                                          Corriente_Calibracion_1,
                                          val_Corriente_Calibracion_1,
                                          Corriente_Calibracion_2,
                                          val_Corriente_Calibracion_2
                                          ]
                                          )

        ################################################################HOJA 3
        hoja_parametros_servicios = wb.create_sheet(title="Servicios")

        ##########ETIQUETAS
        hoja_parametros_servicios['A1'] = "PUNTOS"
        hoja_parametros_servicios['B1'] = "FRESADOS"
        hoja_parametros_servicios['C1'] = "ALARMA PUNTOS"
        hoja_parametros_servicios['D1'] = "ALARMA FRESADOS"
        hoja_parametros_servicios['E1'] = "FUERZA FRESADO"
        hoja_parametros_servicios['F1'] = "FUERZA CAMBIO ELECTRODO"
        hoja_parametros_servicios['G1'] = "ACERCAMIENTO"
        hoja_parametros_servicios['H1'] = "APRIETE"
        hoja_parametros_servicios['I1'] = "PUNTOS POR PIEZA"
        hoja_parametros_servicios['J1'] = "NUMERO CURVA"
        hoja_parametros_servicios['K1'] = "PORCENTAJE DE INCREMENTO"

        Puntos = self.caja_39.value()

        Fresados = self.caja_40.value()

        Alarma_Puntos = self.caja_41.value()
        Alarma_Fresados = self.caja_42.value()

        Fuerza_Fresado = self.caja_46.value()
        Fuerza_Cambio_Electrodo = self.caja_47.value()

        Tiempo_Acercamiento = self.caja_43.value()
        Tiempo_Apriete = self.caja_44.value()

        Puntos_Pieza = self.caja_45.value()

        Numero_Curva = self.caja_38.value()
        Porcentaje = self.caja_37.value()

        hoja_parametros_servicios.append(
                                        [Puntos, Fresados,
                                        Alarma_Puntos, Alarma_Fresados,
                                        Fuerza_Fresado, Fuerza_Cambio_Electrodo,
                                        Tiempo_Acercamiento, Tiempo_Apriete,
                                        Puntos_Pieza, Numero_Curva, Porcentaje
                                        ]
                                        )

        wb.save(direccion)

    def _cargo_excel_(self):
        """
        ESTA FUNCION CARGA LOS PARAMETROS
        DESDE UN ARCHIVO EXCEL.
        """

        filename = QtGui.QFileDialog.getOpenFileName(self,
                                                     "Abrir",
                                                     "/home",
                                                     "*.xls *.xlsx"
                                                    )

        wb2 = load_workbook(filename)

        print (wb2.get_sheet_names())
        print (filename)

    def _abro_dat_(self):
        """
        ABRO UN ARCHIVO *.TXT DESDE LA UBICACION
        SELECCIONADA :)
        """

        try:
            filename = QtGui.QFileDialog.getOpenFileName(self,
                                                         "Abrir",
                                                         "",
                                                         "*.dat"
                                                        )

            self.x = np.loadtxt(filename)

        except:
            pass

        self._bloqueo_signals_('S')

        self._seteo_cajas_parametros_()
        self._oculto_cosas_()

        self._bloqueo_signals_('N')

    def _guardo_en_txt_(self):
        """
        GUARDA EN UN TXT TODOS LOS PARAMETROS
        DEL PROGRAMA.
        """

        import datetime
        fecha = str(datetime.datetime.now().date())
        nombre_1 = ("save_" + fecha + ".txt")
        nombre_2 = ("save_" + fecha + ".dat")

        np.savetxt(nombre_2, self.x, fmt='%.4e')

        fila_maxima = self.caja_26.maximum()

        fo = open(nombre_1, "w")

        for i in range(0, fila_maxima):
            for j in range(0, 21):
                if self.x[i][j] == "]":
                    fo.write("\n")
                else:
                    fo.write(str(self.x[i][j]) + ";")
            fo.write("\n")

        fo.close()

    def _guardo_como_en_txt_(self):
        """
        ESTA FUNCIO GUARDA UN ARCHIVO *.TXT
        EN LA UBICACION SELLECIONADA
        """

        import datetime
        fecha = str(datetime.datetime.now().date())
        nombre_1 = ("save_" + fecha + ".txt")

        direccion = QtGui.QFileDialog.getSaveFileName(self, "hola", nombre_1)

        fila_maxima = self.caja_26.maximum()

        try:
            fo = open(direccion, "w")

            for i in range(0, fila_maxima):
                for j in range(0, 21):
                    if self.x[i][j] == "]":
                        fo.write("\n")
                    else:
                        fo.write(str(self.x[i][j]) + ";")
                fo.write("\n")

            fo.close()

        except:
            pass

    def _guardo_backup_(self):
        """
        """

        #self._selecciono_directorio_backup_()

        import datetime
        fecha = str(datetime.datetime.now().date())
        nombre_1 = ("___BackupControl_" + fecha + ".dat")

        np.savetxt(nombre_1, self.x, fmt='%.4e')

    def _selecciono_directorio_saves_(self):
        """
        """

        directorio = os.getcwd()

        try:
            os.mkdir(directorio + "/saves")

        except:
            pass

        os.chdir(directorio + "/saves")

    def _selecciono_directorio_backup_(self):
        """
        """

        directorio = os.getcwd()

        try:
            os.mkdir(directorio + "/backup")

        except:
            pass

        os.chdir(directorio + "/backup")

    def _seteo_cajas_parametros_(self):
        """
        """

        self._bloqueo_signals_('S')

        fila_actual = self.caja_26.value() - 1

        self.caja_1.setValue(self.x[fila_actual][0])
        self.caja_2.setValue(self.x[fila_actual][1])
        self.caja_3.setValue(self.x[fila_actual][2])

        self.caja_4.setValue(self.x[fila_actual][3])

        #self.caja_5.setValue(self.x[fila_actual][4])  # Intensidad
        self.caja_5.setValue(self._calculo_valor_corriente(self.x[fila_actual][4]))

        self.caja_6.setValue(self.x[fila_actual][5])

        self.caja_7.setValue(self.x[fila_actual][6])
        #self.caja_8.setValue(self.x[fila_actual][7])  # Intensidad
        self.caja_5.setValue(self._calculo_valor_corriente(self.x[fila_actual][7]))
        self.caja_9.setValue(self.x[fila_actual][8])
        self.caja_10.setValue(self.x[fila_actual][9])

        self.caja_27.setValue(self.x[fila_actual][10])
        #self.caja_28.setValue(self.x[fila_actual][11])
        self.caja_28.setValue(self._calculo_valor_corriente(self.x[fila_actual][11]))
        self.caja_29.setValue(self.x[fila_actual][12])
        #self.caja_30.setValue(self.x[fila_actual][13])
        self.caja_30.setValue(self._calculo_valor_corriente(self.x[fila_actual][13]))

        self.caja_11.setValue(self.x[fila_actual][14])
        self.caja_12.setValue(self.x[fila_actual][15])

        self.caja_13.setValue(self.x[fila_actual][16])
        #self.caja_14.setValue(self.x[fila_actual][17])  # Intensidad
        self.caja_5.setValue(self._calculo_valor_corriente(self.x[fila_actual][17]))
        self.caja_15.setValue(self.x[fila_actual][18])

        self.caja_16.setValue(self.x[fila_actual][19])
        #self.caja_18.setValue(self.x[fila_actual][20])  # Fuerza
        self.caja_18.setValue(self._calculo_valor_fuerza(self.x[fila_actual][20]))

        self._calculo_tiempo_proceso()

        self._bloqueo_signals_('N')

    def _seteo_cajas_calibracion_(self):
        """
        """

        self.caja_17.setValue(self.y[0][0])  # Acercamiento
        self.caja_33.setValue(self.y[0][1])  # Apriete
        self.caja_34.setValue(self.y[0][2])  # Presion
        self.caja_35.setValue(self.y[0][3])  # Soldadura
        self.caja_36.setValue(self.y[0][4])  # Mantenido

        self.caja_20.setValue(self.y[0][5])  # Fuerza 1, tension
        self.caja_21.setValue(self.y[0][6])  # Fuerza 2, tension
        #self.caja_23.setValue(self.y[0][7])  # Fuerza 1, daN
        #self.caja_24.setValue(self.y[0][8])  # Fuerza 2, daN

        self.caja_19.setValue(self.y[0][9])  # Corriente 1, %
        self.caja_31.setValue(self.y[0][10])  # Corriente 2, %
        #self.caja_22.setValue(self.y[0][11])  # Corriente 1, KA
        #self.caja_32.setValue(self.y[0][12])  # Corriente 2, KA

    def _seteo_cajas_servicios_(self):
        """
        """

        self.caja_39.setValue(self.z[0][0])  # Puntos
        self.caja_40.setValue(self.z[0][1])  # Fresados

        self.caja_41.setValue(self.z[0][2])  # Alarma Puntos
        self.caja_42.setValue(self.z[0][3])  # Alarma Fresado

        self.caja_46.setValue(self.z[0][4])  # Fuerza Fresado
        self.caja_47.setValue(self.z[0][5])  # Fuerza Cambio Electrodo

        self.caja_43.setValue(self.z[0][6])  # Acercamiento
        self.caja_44.setValue(self.z[0][7])  # Apriete

        self.caja_45.setValue(self.z[0][8])  # Puntos por Pieza

        self.caja_38.setValue(self.z[0][9])  # Numero de Curva
        self.caja_37.setValue(self.z[0][10])  # Porcentaje de Incremento

    def _seteo_cajas_configuracion_(self):
        """
        """

        pass

    def _bloqueo_signals_(self, Modo):
        if Modo == 'S':
            self.caja_1.blockSignals(True)
            self.caja_2.blockSignals(True)
            self.caja_3.blockSignals(True)
            self.caja_4.blockSignals(True)
            self.caja_5.blockSignals(True)
            self.caja_6.blockSignals(True)
            self.caja_7.blockSignals(True)
            self.caja_8.blockSignals(True)
            self.caja_9.blockSignals(True)
            self.caja_10.blockSignals(True)
            self.caja_11.blockSignals(True)
            self.caja_12.blockSignals(True)
            self.caja_27.blockSignals(True)
            self.caja_28.blockSignals(True)
            self.caja_29.blockSignals(True)
            self.caja_30.blockSignals(True)
            self.caja_13.blockSignals(True)
            self.caja_14.blockSignals(True)
            self.caja_15.blockSignals(True)
            self.caja_16.blockSignals(True)
            self.caja_18.blockSignals(True)
        else:
            self.caja_1.blockSignals(False)
            self.caja_2.blockSignals(False)
            self.caja_3.blockSignals(False)
            self.caja_4.blockSignals(False)
            self.caja_5.blockSignals(False)
            self.caja_6.blockSignals(False)
            self.caja_7.blockSignals(False)
            self.caja_8.blockSignals(False)
            self.caja_9.blockSignals(False)
            self.caja_10.blockSignals(False)
            self.caja_11.blockSignals(False)
            self.caja_12.blockSignals(False)
            self.caja_27.blockSignals(False)
            self.caja_28.blockSignals(False)
            self.caja_29.blockSignals(False)
            self.caja_30.blockSignals(False)
            self.caja_13.blockSignals(False)
            self.caja_14.blockSignals(False)
            self.caja_15.blockSignals(False)
            self.caja_16.blockSignals(False)
            self.caja_18.blockSignals(False)

        self._oculto_cosas_()

    def _nuevo_(self):
        #QtGui.QMessageBox.about(self, "hola", "nene")
        QtGui.QMessageBox.information(self, "Alerta", "funcion no disponible")

    def _copiar_programas_(self):
        QtGui.QMessageBox.information(self, "Alerta", "funcion no disponible")

    def _configuracion_(self):
        QtGui.QMessageBox.information(self, "Alerta", "funcion no disponible")

    def _manual_(self):
        QtGui.QMessageBox.information(self, "Alerta", "funcion no disponible")

    def _acerca_de_(self):
        texto_menu = "Acerca de"

        texto_info = """
        Este software permite manipular el control de soldadura
        y se distribuye de forma gratuita bajo licencia GNU GPL.
                     """

        texto_version = "\n VERSION: 1.2.1"
        texto_contacto = "\n CONTACTO: fe.mas.ingenieria@gmail.com"

        QtGui.QMessageBox.about(self, texto_menu, texto_info
                                + texto_version
                                + texto_contacto
                                )

    def _acerca_de_QT_(self):
        texto_menu = "Acerca de QT"

        QtGui.QMessageBox.aboutQt(self, texto_menu)

    def _addmpl_(self, fig):
        self.canvas = FigureCanvas(fig)
        self.gridLayout_2.addWidget(self.canvas, 3, 0)
        self.canvas.draw()
        self.toolbar = NavigationToolbar(self.canvas,
                                         self,
                                         coordinates=True
                                         )
        self.gridLayout_2.addWidget(self.toolbar)


def _cargo_grafica_():
    fig1 = Figure(facecolor='w', tight_layout='rect')

    ax1f1 = fig1.add_subplot(111)

    ax1f1.axhline(y=.4, label="Soldadura", linestyle='--', linewidth=2.5, color='g')
    ax1f1.axhline(y=.5, label="% Incremento", linestyle='--', linewidth=2.5, color='r')
    ax1f1.axvline(x=.4, label="Puntos", linestyle='--', linewidth=2.5, color='b')

    ax1f1.legend()
    ax1f1.grid(True)

    window._addmpl_(fig1)

if __name__ == "__main__":
    app = QtGui.QApplication(sys.argv)
    window = MyApp()

    _cargo_grafica_()

    window.show()
    sys.exit(app.exec_())